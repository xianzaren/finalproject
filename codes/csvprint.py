import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors

# colored Excel
def save_colored_excel(data, filepath, colormap_name):
    # Normalize the data to [0, 1]
    data_normalized = (data - data.min()) / (data.max() - data.min())

    #colormap
    cmap = plt.get_cmap(colormap_name)
    norm = mcolors.Normalize(vmin=0, vmax=1)

    #Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Colored Data"

    # Write data with dynamic colormap-based coloring
    for i, row in enumerate(data_normalized, start=1):
        for j, value in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=value)

            # Convert value to RGB color using the colormap
            rgba = cmap(norm(value))
            rgb = tuple(int(255 * c) for c in rgba[:3])  # Extract RGB from RGBA
            hex_color = "{:02X}{:02X}{:02X}".format(*rgb)  # Convert to HEX

            # Apply the color as cell background
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    # Save the workbook
    wb.save(filepath)
    print(f"Excel file saved to: {filepath}")

"""
parameter you need to change
"""
shape = (82, 63)
res = (4, 4)
octaves = 5
colormap_name='hot'
prepath = f'../color\\{colormap_name}'

# Filepaths
csv_filepath = f"{prepath}\\{shape}_{res}_{octaves}.csv"
xlsx_filepath = f"{prepath}\\{shape}_{res}_{octaves}_{colormap_name}_colored.xlsx"

# Load the CSV data
data = np.loadtxt(csv_filepath, delimiter=",", skiprows=1)  # Skip the header row

# Save the data as a colored Excel file
save_colored_excel(data, xlsx_filepath, colormap_name)
