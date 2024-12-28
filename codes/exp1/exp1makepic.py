import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from scipy.fftpack import fft2, fftshift
import os

# 1) 生成 Perlin 噪声的函数（与原始类似）
def generate_seamless_perlin_noise_2d(shape, res, octaves=5, persistence=0.5):
    def f(t):
        return 6 * t**5 - 15 * t**4 + 10 * t**3

    def gradient_grid(res):
        angles = 2 * np.pi * np.random.rand(res[0], res[1])
        gradients = np.dstack((np.cos(angles), np.sin(angles)))
        return gradients

    def perlin(x, y, gradients):
        x0 = np.floor(x).astype(int) % gradients.shape[0]
        y0 = np.floor(y).astype(int) % gradients.shape[1]
        x1 = (x0 + 1) % gradients.shape[0]
        y1 = (y0 + 1) % gradients.shape[1]

        dx = x - x0
        dy = y - y0
        sx = f(dx)
        sy = f(dy)

        g00 = gradients[x0, y0]
        g10 = gradients[x1, y0]
        g01 = gradients[x0, y1]
        g11 = gradients[x1, y1]

        n00 = g00[..., 0] * dx + g00[..., 1] * dy
        n10 = g10[..., 0] * (dx - 1) + g10[..., 1] * dy
        n01 = g01[..., 0] * dx + g01[..., 1] * (dy - 1)
        n11 = g11[..., 0] * (dx - 1) + g11[..., 1] * (dy - 1)

        nx0 = (1 - sx) * n00 + sx * n10
        nx1 = (1 - sx) * n01 + sx * n11
        return (1 - sy) * nx0 + sy * nx1

    noise = np.zeros(shape)
    frequency = 1
    amplitude = 1
    max_amplitude = 0

    for _ in range(octaves):
        res_x, res_y = int(res[0] * frequency), int(res[1] * frequency)
        gradients = gradient_grid((res_x, res_y))

        delta = (res_x / shape[0], res_y / shape[1])
        grid_x, grid_y = np.meshgrid(
            np.arange(0, shape[1]) * delta[1],
            np.arange(0, shape[0]) * delta[0]
        )

        noise += amplitude * perlin(grid_x, grid_y, gradients)
        max_amplitude += amplitude
        amplitude *= persistence
        frequency *= 2

    return noise / max_amplitude


# 2) FFT -> 计算功率谱（可根据需要保留或去除）
def compute_power_spectrum(noise):
    fft_values = fftshift(fft2(noise))
    magnitude = np.abs(fft_values)
    power_spectrum = np.log1p(magnitude ** 2)
    return power_spectrum


# 3) 保存带颜色的 Excel
def save_colored_excel(noise, filepath, colormap):
    # 归一化到 [0,1]
    noise = (noise - noise.min()) / (noise.max() - noise.min())
    norm = mcolors.Normalize(vmin=0, vmax=1)
    cmap = plt.get_cmap(colormap)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Elevation Data"

    for i, row in enumerate(noise, start=1):
        for j, value in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            rgba = cmap(norm(value))
            rgb = tuple(int(255 * c) for c in rgba[:3])  # (R,G,B)
            hex_color = "{:02X}{:02X}{:02X}".format(*rgb)
            cell.fill = PatternFill(
                start_color=hex_color,
                end_color=hex_color,
                fill_type="solid"
            )

    wb.save(filepath)


# 4) 生成不带坐标轴的纯粹标量场图
def plot_scalar_field(shape, noise, colormap, filepath):
    dpi = 100
    figsize_width = shape[1] / dpi + 1e-9
    figsize_height = shape[0] / dpi + 1e-9

    plt.figure(figsize=(figsize_width, figsize_height), dpi=dpi)
    plt.imshow(noise, cmap=colormap, origin='upper')
    plt.axis('off')
    plt.gca().set_axis_off()
    plt.subplots_adjust(top=1, bottom=0, left=0, right=1, hspace=0, wspace=0)
    plt.margins(0, 0)
    plt.savefig(filepath, dpi=dpi, bbox_inches='tight', pad_inches=0)
    plt.close()
    print(f"纯粹标量场图像已保存为: {filepath}")


def main():
    """
    逻辑：
    1) 先生成多份噪声数据 (这里用 5 种 res 举例)。
    2) 不同的噪声数据都保存到 uncolor 文件夹 (CSV)。
    3) 对每份噪声使用 3 种 colormap 做可视化：
       a) 带 colorbar 的图 -> 保存到对应 colormap 的文件夹
       b) 不带坐标轴的图 -> 也保存到对应 colormap 的文件夹
       c) 染色 Excel -> 保存到 colored 文件夹
    """

    # --------- 基本参数 ---------
    shape = (630, 820)       # 图像大小
    octaves = 5              # Perlin noise 的叠加层数
    res_list = [(1, 1), (3, 3), (5, 5), (7, 7), (9, 9)]  # 5 种不同频率或分辨率
    colormap_list = ['gray', 'rainbow', 'hot']          # 3 种 colormap

    # --------- 输出文件夹路径 ---------
    uncolor_dir = r'../../data/uncolor'
    colored_dir = r'../../data/colored'
    # 根据 colormap 不同，图像将保存到不同子目录
    image_dir_map = {
        'gray':    r'../../color/gray',
        'rainbow': r'../../color/rainbow',
        'hot':     r'../../color/hot'
    }

    # 若目录不存在，创建它
    os.makedirs(uncolor_dir, exist_ok=True)
    os.makedirs(colored_dir, exist_ok=True)
    for cdir in image_dir_map.values():
        os.makedirs(cdir, exist_ok=True)

    # --------- 主循环：遍历 5 种不同的 res 生成数据 ---------
    for res in res_list:
        print(f"正在生成 Perlin 噪声: shape={shape}, res={res}, octaves={octaves}")
        noise = generate_seamless_perlin_noise_2d(shape, res, octaves)

        # 先整体归一化（有助于后续可视化）
        noise = (noise - noise.min()) / (noise.max() - noise.min())

        # 1) 保存 CSV（含噪声数据）到 uncolor 文件夹
        csv_filename = f"Noise_{shape}_{res}_{octaves}.csv"
        csv_filepath = os.path.join(uncolor_dir, csv_filename)
        np.savetxt(csv_filepath, noise, delimiter=",", header="Normalized Elevation Data", comments="")
        print(f"CSV 已保存(未染色数据): {csv_filepath}")

        # 2) 对此噪声使用三种不同的 colormap 可视化
        for cm in colormap_list:
            print(f"  使用 colormap={cm} 绘图并保存...")

            # (a) 带 colorbar 的图
            plt.figure(figsize=(8, 6))
            plt.imshow(noise, cmap=cm, origin='upper')
            plt.colorbar(label='Elevation')
            plt.title(f"Perlin Noise (res={res}, colormap={cm})")
            plt.axis('off')
            # 保存到对应 colormap 的文件夹
            colorbar_png_name = f"Colorbar_{cm}_{shape}_{res}_{octaves}.png"
            colorbar_png_path = os.path.join(image_dir_map[cm], colorbar_png_name)
            plt.savefig(colorbar_png_path, dpi=150, bbox_inches='tight', pad_inches=0)
            plt.close()
            print(f"    带 colormap 的图像已保存到: {colorbar_png_path}")

            # (b) 纯粹标量场图，不带坐标轴
            scalar_field_png_name = f"ScalarField_{cm}_{shape}_{res}_{octaves}.png"
            scalar_field_png_path = os.path.join(image_dir_map[cm], scalar_field_png_name)
            plot_scalar_field(shape, noise, cm, scalar_field_png_path)

            # (c) Excel 带颜色 -> 保存到 colored 文件夹
            excel_filename = f"Colored_{cm}_{shape}_{res}_{octaves}.xlsx"
            excel_filepath = os.path.join(colored_dir, excel_filename)
            save_colored_excel(noise, excel_filepath, cm)
            print(f"    带颜色的 Excel 已保存到: {excel_filepath}")

        print("-----")

if __name__ == "__main__":
    main()
