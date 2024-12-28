import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from scipy.fftpack import fft2, fftshift
import os

#Perlin noise
def generate_seamless_perlin_noise_2d(shape, res, octaves=5, persistence=0.5):
    def f(t):
        return 6 * t**5 - 15 * t**4 + 10 * t**3

    def gradient_grid(res):
        angles = 2 * np.pi * np.random.rand(res[0], res[1])
        gradients = np.dstack((np.cos(angles), np.sin(angles)))
        return gradients

    def perlin(x, y, gradients):
        x0 = np.floor(x).astype(int) % gradients.shape[0]
        y0 = np.floor(y).astype(int) % gradients.shape[1]
        x1 = (x0 + 1) % gradients.shape[0]
        y1 = (y0 + 1) % gradients.shape[1]

        dx = x - x0
        dy = y - y0
        sx = f(dx)
        sy = f(dy)

        g00 = gradients[x0, y0]
        g10 = gradients[x1, y0]
        g01 = gradients[x0, y1]
        g11 = gradients[x1, y1]

        n00 = g00[..., 0] * dx + g00[..., 1] * dy
        n10 = g10[..., 0] * (dx - 1) + g10[..., 1] * dy
        n01 = g01[..., 0] * dx + g01[..., 1] * (dy - 1)
        n11 = g11[..., 0] * (dx - 1) + g11[..., 1] * (dy - 1)

        nx0 = (1 - sx) * n00 + sx * n10
        nx1 = (1 - sx) * n01 + sx * n11
        return (1 - sy) * nx0 + sy * nx1

    noise = np.zeros(shape)
    frequency = 1
    amplitude = 1
    max_amplitude = 0

    for _ in range(octaves):
        res_x, res_y = int(res[0] * frequency), int(res[1] * frequency)
        gradients = gradient_grid((res_x, res_y))

        delta = (res_x / shape[0], res_y / shape[1])
        grid_x, grid_y = np.meshgrid(
            np.arange(0, shape[1]) * delta[1],
            np.arange(0, shape[0]) * delta[0]
        )

        noise += amplitude * perlin(grid_x, grid_y, gradients)
        max_amplitude += amplitude
        amplitude *= persistence
        frequency *= 2

    return noise / max_amplitude

# FFT
def compute_power_spectrum(noise):
    fft_values = fftshift(fft2(noise))
    magnitude = np.abs(fft_values)
    power_spectrum = np.log1p(magnitude ** 2)  # Logarithmic scale for better visualization
    return power_spectrum

# colored Excel
def save_colored_excel(noise, filepath, colormap):

    # Normalize
    noise = (noise - noise.min()) / (noise.max() - noise.min())
    norm = mcolors.Normalize(vmin=0, vmax=1)

    #colormap
    cmap = plt.get_cmap(colormap)


    # Create an Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Elevation Data"

    # Write data with dynamic colormap-based coloring
    for i, row in enumerate(noise, start=1):
        for j, value in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=value)

            # Convert value to RGB color using the colormap
            rgba = cmap(norm(value))
            rgb = tuple(int(255 * c) for c in rgba[:3])  # Extract RGB from RGBA
            hex_color = "{:02X}{:02X}{:02X}".format(*rgb)  # Convert to HEX

            # Apply the color as cell background
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    # Save the workbook
    wb.save(filepath)

def plot_scalar_field(shape, noise, colormap, filepath):
    dpi = 100  # 设置 DPI
    figsize_width = shape[1] / dpi + 1e-9  # 增加微小偏移，避免浮点数舍入误差
    figsize_height = shape[0] / dpi + 1e-9

    plt.figure(figsize=(figsize_width, figsize_height), dpi=dpi)
    plt.imshow(noise, cmap=colormap, origin='upper')
    plt.axis('off')
    plt.gca().set_axis_off()
    plt.subplots_adjust(top=1, bottom=0, left=0, right=1, hspace=0, wspace=0)
    plt.margins(0, 0)
    plt.savefig(filepath, dpi=dpi, bbox_inches='tight', pad_inches=0)
    plt.close()
    print(f"标量场图像已保存为: {filepath}")

def main():
    colormap_name = 'coolwarm'
    shape = (820, 630)
    octaves = 5
    color_list = ['rainbow','gray','hot']
    res_list = [(1, 1), (3, 3), (5, 5), (7, 7), (9, 9)]  # 基础网格分辨率
    for res in res_list:
        prepath = f'../color\\{colormap_name}\\{shape}{res}{octaves}'
        os.makedirs(prepath, exist_ok=True)
        print(f"正在生成: res={res}, octaves={octaves}")
        #Perlin noise
        noise = generate_seamless_perlin_noise_2d(shape, res, octaves)

        # Normalize
        noise = (noise - noise.min()) / (noise.max() - noise.min())

        # 保存excel
        np.savetxt(f"{prepath}\\{shape}_{res}_{octaves}.csv", noise, delimiter=",", header="Normalized Elevation Data", comments="")
        print("海拔数据已保存为 csv文件")

        # excel染色
        excel_filepath = f"{prepath}\\{shape}_{res}_{octaves}_{colormap_name}_colored.xlsx"
        save_colored_excel(noise, excel_filepath, colormap_name)
        print(f"带颜色的Excel文件已保存为: {excel_filepath}")

        colormap = plt.get_cmap(colormap_name)

        # 计算power spectrum
        power_spectrum = compute_power_spectrum(noise)

        # 生成图with colormap
        plt.figure(figsize=(8, 6))
        plt.imshow(noise, cmap=colormap, origin='upper')
        plt.colorbar(label='Elevation')
        plt.title(f"Seamless Perlin Noise ({colormap_name.capitalize()} Colormap)")
        plt.axis('off')
        plt.savefig(f"{prepath}\\seamless_terrain_{colormap_name}_{shape}_{res}_{octaves}.png", dpi=300, bbox_inches=None, pad_inches=0)
        # plt.show()
        # 生成图without colormap
        scalar_field_filepath = f"{prepath}\\scalar_field_{colormap_name}_{shape}_{res}_{octaves}.png"
        plot_scalar_field(shape, noise, colormap_name, scalar_field_filepath)

if __name__ == "__main__":
    main()
